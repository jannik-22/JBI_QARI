from pathlib import Path
from typing import List
import fitz  # PyMuPDF
import pandas as pd
from schema import JBIEntry
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def extract_text_from_pdf(pdf_path: Path) -> str:
    with fitz.open(pdf_path) as doc:
        return "\n".join(page.get_text() for page in doc)

def save_results_to_excel(data, output_path: str):
    # DataFrame erstellen und Spalten so anordnen, dass 'ID' ganz links steht
    df = pd.DataFrame(data)
    if "ID" in df.columns:
        cols = ["ID"] + [c for c in df.columns if c != "ID"]
        df = df[cols]

    df.to_excel(output_path, index=False)

    # Excel mit openpyxl laden und anpassen
    workbook = load_workbook(output_path)
    sheet = workbook.active

    # Spaltenbreite setzen und Zeilenumbruch aktivieren
    for i, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(i)
        sheet.column_dimensions[col_letter].width = 40  # Spaltenbreite

        for cell in sheet[col_letter]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")  # Zeilenumbruch + oben ausrichten

    workbook.save(output_path)


def save_multi_sheet_excel(data_dict: dict[str, list[dict]], output_path: str):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, data in data_dict.items():
            df = pd.DataFrame(data)
            # Spalten so anordnen, dass 'ID' ganz links steht
            if "ID" in df.columns:
                cols = ["ID"] + [c for c in df.columns if c != "ID"]
                df = df[cols]

            df.to_excel(writer, index=False, sheet_name=sheet_name[:31])  # Excel max. 31 Zeichen

    # Formatierung (automatische Spaltenbreite, Zeilenumbruch)
    workbook = load_workbook(output_path)
    for sheet in workbook.worksheets:
        for i, col in enumerate(sheet.columns, 1):
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            sheet.column_dimensions[get_column_letter(i)].width = min(max_len + 5, 50)
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(output_path)